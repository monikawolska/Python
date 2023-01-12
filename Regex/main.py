import datetime
import openpyxl
import miary
import pandas as pd

# Important!
# Aby kod zadziałał należy zainstalować starą wersje openpyxl
# Komendy do anacondy:
# pip uninstall openpyxl
# pip install openpyxl==2.6.3

data = str(datetime.datetime.now())

nazwa = 'Kategoryzacja_' + data[:4] + data[5:7] + data[8:10] + '.xlsm'
path = r'K:\DHQ\Groups\Production\Pion Logistyki\Dzial Danych Podstawowych\Biblia\BAZA kategoryzaji\Python_dodawanie miar'
kategoryzacja = path + '\\' + nazwa

wb = openpyxl.load_workbook(filename = kategoryzacja, read_only = False, keep_vba = True)
sheet = wb['PKWiU']
df = pd.read_excel(kategoryzacja, sheet_name='PKWiU')

for index, row in df.iterrows():
    result = miary.miary(miary.preparing(row[2]))
    try:
        sheet.cell(index + 2, 17).value = float(result[0])
    except:
        sheet.cell(index + 2, 17).value = result[0]
    sheet.cell(index + 2, 18).value = result[1]
    sheet.cell(index + 2, 19).value = result[2]

# saving excel
wb.save(kategoryzacja)

print("Zakończono")
# !Przy wykonywaniu programu pojawi się UserWarning - należy to zignorować